from __future__ import annotations

import csv
import io
import json
import os
import re
import subprocess
import tempfile
from dataclasses import replace
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from collections.abc import Mapping
from typing import Any, Callable, cast

from agents.document_ai.schemas import (
    ExtractedRecord,
    ExtractionContext,
    ExtractionResult,
    SourceLocation,
)
from agents.shared.schemas import Confidence, Evidence, Money


SCHEMA_VERSION = "document-extraction.v1"

_HEADER_ALIASES = {
    "date": "date",
    "transaction_date": "date",
    "posting_date": "date",
    "invoice_date": "date",
    "payment_date": "date",
    "receipt_date": "date",
    "claim_date": "date",
    "reference": "reference",
    "ref": "reference",
    "transaction_reference": "reference",
    "transaction_id": "reference",
    "invoice_number": "reference",
    "invoice_no": "reference",
    "bill_number": "reference",
    "receipt_number": "reference",
    "payment_reference": "reference",
    "contract_number": "reference",
    "policy_number": "policy_number",
    "policy_no": "policy_number",
    "claim_number": "claim_reference",
    "claim_ref": "claim_reference",
    "amount": "amount",
    "total": "amount",
    "total_amount": "amount",
    "gross_amount": "amount",
    "premium_amount": "amount",
    "claim_amount": "amount",
    "currency": "currency",
    "ccy": "currency",
    "type": "type",
    "record_type": "type",
    "category": "type",
    "party": "party_name",
    "party_name": "party_name",
    "counterparty": "party_name",
    "supplier": "party_name",
    "vendor": "party_name",
    "customer": "party_name",
    "broker": "party_name",
    "payer": "party_name",
    "payee": "party_name",
    "insured": "party_name",
    "description": "description",
    "memo": "description",
    "narrative": "description",
    "due_date": "due_date",
    "start_date": "start_date",
    "end_date": "end_date",
    "status": "status",
}

_REQUIRED_FIELDS = {
    "transaction": ("date", "reference", "amount", "currency"),
    "payment": ("date", "reference", "amount", "currency"),
    "invoice": ("date", "reference", "amount", "currency"),
    "bill": ("date", "reference", "amount", "currency"),
    "receipt": ("date", "reference", "amount", "currency"),
    "premium": ("date", "reference", "amount", "currency"),
    "claim": ("date", "reference", "amount", "currency"),
    "contract": ("date", "reference", "party_name"),
    "external_party": ("party_name",),
}

_RECORD_TYPES = set(_REQUIRED_FIELDS)
_DATE_FIELDS = {"date", "due_date", "start_date", "end_date"}


def extract_document(
    content: bytes,
    context: ExtractionContext,
    ocr_backend: Callable[[object, str], str] | None = None,
    ocr_language: str = "eng",
    pdf_page_renderer: Callable[[bytes, int], object] | None = None,
) -> ExtractionResult:
    context.validate()
    if not content:
        return _empty_result(context, "empty-document", "document content is empty")

    suffix = Path(context.file_name).suffix.lower()
    content_type = context.content_type.lower()
    if suffix == ".csv" or content_type in {"text/csv", "application/csv"}:
        return _extract_csv_bytes(content, context)
    if suffix in {".xlsx", ".xlsm"} or "spreadsheetml" in content_type:
        return _extract_excel_bytes(content, context)
    if suffix == ".pdf" or content_type == "application/pdf":
        return _extract_pdf_bytes(
            content,
            context,
            ocr_backend=ocr_backend,
            ocr_language=ocr_language,
            page_renderer=pdf_page_renderer,
        )
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"} or content_type.startswith("image/"):
        return _extract_image_bytes(content, context, ocr_backend, ocr_language)
    return _empty_result(context, "unsupported-format", f"unsupported document format: {suffix or content_type}")


def extract_csv(path: str | Path, source_document_id: str) -> list[ExtractedRecord]:
    source_path = Path(path)
    context = ExtractionContext(
        organization_id="legacy",
        source_document_id=source_document_id,
        ingestion_batch_id="legacy-batch",
        extraction_version_id="legacy-version",
        file_name=source_path.name,
        content_type="text/csv",
    )
    return list(extract_document(source_path.read_bytes(), context).records)


def _extract_csv_bytes(content: bytes, context: ExtractionContext) -> ExtractionResult:
    text = _decode_text(content)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return _empty_result(context, "csv", "CSV header row is missing")
    records = []
    for ordinal, row in enumerate(reader, start=1):
        records.append(
            _record_from_row(
                row,
                context,
                source_record_id=f"row-{ordinal}",
                source_location=SourceLocation(row_number=ordinal + 1),
                parser="csv",
                base_confidence=0.98,
            )
        )
    return _finalize_result(context, "csv", records)


def _extract_excel_bytes(content: bytes, context: ExtractionContext) -> ExtractionResult:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        return _empty_result(context, "excel", f"Excel parser unavailable: {exc}")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl exposes several format-specific exceptions
        return _empty_result(context, "excel", f"invalid Excel workbook: {exc}")

    records: list[ExtractedRecord] = []
    warnings: list[str] = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            warnings.append(f"empty-sheet:{sheet.title}")
            continue
        header_values = [str(value).strip() if value is not None else "" for value in headers]
        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            row = {
                header_values[index]: _cell_text(value)
                for index, value in enumerate(values)
                if index < len(header_values) and header_values[index]
            }
            records.append(
                _record_from_row(
                    row,
                    context,
                    source_record_id=f"{_slug(sheet.title)}-row-{row_number - 1}",
                    source_location=SourceLocation(row_number=row_number, sheet_name=sheet.title),
                    parser="excel",
                    base_confidence=0.98,
                )
            )
    result = _finalize_result(context, "excel", records, warnings)
    return replace(result, metadata={"sheet_count": str(len(workbook.sheetnames))})


def _extract_pdf_bytes(
    content: bytes,
    context: ExtractionContext,
    ocr_backend: Callable[[object, str], str] | None,
    ocr_language: str,
    page_renderer: Callable[[bytes, int], object] | None,
) -> ExtractionResult:
    try:
        try:
            from pypdf import PdfReader

            reader_type: Any = PdfReader
        except ImportError:
            from PyPDF2 import PdfReader as LegacyPdfReader

            reader_type = LegacyPdfReader
        reader = reader_type(io.BytesIO(content))
    except Exception as exc:
        return _empty_result(context, "pdf", f"invalid PDF: {exc}")

    records: list[ExtractedRecord] = []
    warnings: list[str] = []
    for page_number, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        parser = "pdf-text"
        base_confidence = 0.90
        if not text.strip():
            renderer = page_renderer or _pymupdf_render_page
            backend = ocr_backend or _pytesseract_ocr
            try:
                image = renderer(content, page_number)
                text = backend(image, ocr_language)
                parser = "pdf-ocr"
                base_confidence = 0.78
                warnings.append(f"page-ocr-fallback:{page_number}")
            except Exception as exc:
                warnings.append(f"page-without-text:{page_number}")
                warnings.append(f"page-ocr-unavailable:{page_number}:{exc}")
                continue
            if not text.strip():
                warnings.append(f"page-ocr-empty:{page_number}")
                continue
        records.append(
            _record_from_text(
                text,
                context,
                source_record_id=f"page-{page_number}",
                source_location=SourceLocation(page_number=page_number),
                parser=parser,
                base_confidence=base_confidence,
            )
        )
    result = _finalize_result(context, "pdf", records, warnings)
    return replace(result, metadata={"page_count": str(len(reader.pages))})


def _pymupdf_render_page(content: bytes, page_number: int) -> object:
    import pymupdf  # type: ignore[import-not-found]
    from PIL import Image

    document = pymupdf.open(stream=content, filetype="pdf")
    try:
        page = document.load_page(page_number - 1)
        pixmap = page.get_pixmap(matrix=pymupdf.Matrix(2, 2), alpha=False)
        return Image.frombytes("RGB", (pixmap.width, pixmap.height), pixmap.samples)
    finally:
        document.close()


def _extract_image_bytes(
    content: bytes,
    context: ExtractionContext,
    ocr_backend: Callable[[object, str], str] | None,
    ocr_language: str,
) -> ExtractionResult:
    try:
        from PIL import Image

        candidate = Image.open(io.BytesIO(content))
        candidate.verify()
        image = Image.open(io.BytesIO(content)).convert("RGB")
    except Exception as exc:
        return _empty_result(context, "image-ocr", f"invalid image: {exc}")

    backend = ocr_backend or _local_ocr
    try:
        text = backend(image, ocr_language)
    except Exception as exc:
        return _empty_result(context, "image-ocr", f"OCR unavailable: {exc}")
    if not text.strip():
        return _empty_result(context, "image-ocr", "OCR returned no text")

    record = _record_from_text(
        text,
        context,
        source_record_id="image-1",
        source_location=SourceLocation(page_number=1),
        parser="image-ocr",
        base_confidence=0.78,
    )
    return _finalize_result(context, "image-ocr", [record])


def _pytesseract_ocr(image: object, language: str) -> str:
    import pytesseract  # type: ignore[import-untyped]
    from PIL import Image, ImageOps

    prepared = ImageOps.autocontrast(ImageOps.grayscale(cast(Image.Image, image)))
    prepared = prepared.resize((prepared.width * 2, prepared.height * 2))
    return pytesseract.image_to_string(prepared, lang=language, config="--psm 6")


def _local_ocr(image: object, language: str) -> str:
    try:
        return _pytesseract_ocr(image, language)
    except Exception:
        if os.name != "nt":
            raise
        return _windows_ocr(image, language)


def _windows_ocr(image: object, language: str) -> str:
    from PIL import Image

    script = r"""
$ErrorActionPreference='Stop'
Add-Type -AssemblyName System.Runtime.WindowsRuntime
$generic=([System.WindowsRuntimeSystemExtensions].GetMethods() | Where-Object {
  $_.Name -eq 'AsTask' -and $_.IsGenericMethod -and $_.GetParameters().Count -eq 1
})[0]
function Await($operation,[Type]$type) {
  $task=$generic.MakeGenericMethod($type).Invoke($null,@($operation)); $task.Wait(); $task.Result
}
[Windows.Storage.StorageFile,Windows.Storage,ContentType=WindowsRuntime] | Out-Null
[Windows.Storage.Streams.IRandomAccessStream,Windows.Storage.Streams,ContentType=WindowsRuntime] | Out-Null
[Windows.Graphics.Imaging.BitmapDecoder,Windows.Graphics.Imaging,ContentType=WindowsRuntime] | Out-Null
[Windows.Graphics.Imaging.SoftwareBitmap,Windows.Graphics.Imaging,ContentType=WindowsRuntime] | Out-Null
[Windows.Media.Ocr.OcrEngine,Windows.Foundation,ContentType=WindowsRuntime] | Out-Null
$file=Await ([Windows.Storage.StorageFile]::GetFileFromPathAsync($env:KORA_OCR_FILE)) ([Windows.Storage.StorageFile])
$stream=Await ($file.OpenAsync([Windows.Storage.FileAccessMode]::Read)) ([Windows.Storage.Streams.IRandomAccessStream])
$decoder=Await ([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream)) ([Windows.Graphics.Imaging.BitmapDecoder])
$bitmap=Await ($decoder.GetSoftwareBitmapAsync()) ([Windows.Graphics.Imaging.SoftwareBitmap])
$engine=[Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages()
if ($null -eq $engine) { throw 'Windows OCR engine is unavailable' }
$result=Await ($engine.RecognizeAsync($bitmap)) ([Windows.Media.Ocr.OcrResult])
$result.Lines | ForEach-Object {
  $first=@($_.Words)[0]
  [pscustomobject]@{Text=$_.Text;X=$first.BoundingRect.X;Y=$first.BoundingRect.Y}
} | ConvertTo-Json -Compress
"""
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temporary:
        path = temporary.name
    try:
        cast(Image.Image, image).save(path, format="PNG")
        environment = dict(os.environ)
        environment["KORA_OCR_FILE"] = path
        completed = subprocess.run(
            ["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True,
            check=True,
            encoding="utf-8",
            errors="replace",
            env=environment,
            timeout=45,
        )
        payload = completed.stdout.strip()
        if not payload:
            raise RuntimeError("Windows OCR returned no text")
        lines = json.loads(payload)
        if isinstance(lines, dict):
            lines = [lines]
        return _order_ocr_lines(lines)
    finally:
        try:
            os.remove(path)
        except FileNotFoundError:
            pass


def _order_ocr_lines(lines: list[dict[str, object]]) -> str:
    ordered = sorted(lines, key=lambda line: (float(str(line["Y"])), float(str(line["X"]))))
    bands: list[list[dict[str, object]]] = []
    for line in ordered:
        if not bands or abs(
            float(str(line["Y"])) - float(str(bands[-1][0]["Y"]))
        ) > 8:
            bands.append([line])
        else:
            bands[-1].append(line)
    return "\n".join(
        "\t".join(
            str(line["Text"])
            for line in sorted(band, key=lambda item: float(str(item["X"])))
        )
        for band in bands
    )


def _record_from_text(
    text: str,
    context: ExtractionContext,
    source_record_id: str,
    source_location: SourceLocation,
    parser: str,
    base_confidence: float,
) -> ExtractedRecord:
    row: dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        match = re.match(r"^([^:=]{2,40})\s*[:=]\s*(.+)$", line)
        if match:
            row[match.group(1).strip()] = match.group(2).strip()
    if not row:
        row["description"] = " ".join(text.split())
    return _record_from_row(row, context, source_record_id, source_location, parser, base_confidence)


def _record_from_row(
    row: Mapping[str, object],
    context: ExtractionContext,
    source_record_id: str,
    source_location: SourceLocation,
    parser: str,
    base_confidence: float,
) -> ExtractedRecord:
    fields: dict[str, str] = {}
    field_confidences: dict[str, float] = {}
    warnings: list[str] = []

    for raw_key, raw_value in row.items():
        key = _normalize_header(str(raw_key))
        canonical = _HEADER_ALIASES.get(key, key)
        value = _cell_text(raw_value).strip()
        if not value:
            continue
        confidence = base_confidence if canonical == key else max(0.0, base_confidence - 0.03)
        if canonical in _DATE_FIELDS:
            value, date_warning = _normalize_date(value)
            if date_warning:
                warnings.append(f"{date_warning}:{canonical}")
                confidence = min(confidence, 0.60)
        elif canonical == "amount":
            value, amount_warning = _normalize_amount(value)
            if amount_warning:
                warnings.append(amount_warning)
                confidence = min(confidence, 0.40)
        elif canonical == "currency":
            value = value.upper()
            if not re.fullmatch(r"[A-Z]{3}", value):
                warnings.append("invalid-currency")
                confidence = min(confidence, 0.50)
        fields[canonical] = value
        field_confidences[canonical] = max(field_confidences.get(canonical, 0), confidence)

    record_type = _infer_record_type(fields, context.file_name)
    fields["type"] = record_type
    field_confidences["type"] = field_confidences.get("type", max(0.60, base_confidence - 0.08))
    required = _REQUIRED_FIELDS.get(record_type, ("date", "reference"))
    missing_fields = tuple(name for name in required if not fields.get(name))
    warnings.extend(f"missing-field:{name}" for name in missing_fields)

    scores = [field_confidences.get(name, 0.0) for name in required]
    confidence_score = sum(scores) / len(scores) if scores else base_confidence
    if missing_fields:
        confidence_score = min(confidence_score, 0.60)
    if warnings:
        confidence_score = min(confidence_score, 0.68)
    confidence_score = round(max(0.0, min(confidence_score, 1.0)), 4)
    quality_flags = _quality_flags(confidence_score, missing_fields, warnings)

    amount = _money_from_fields(fields)
    evidence = Evidence(
        source_document_id=context.source_document_id,
        source_record_id=source_record_id,
        transaction_reference=fields.get("reference", ""),
        occurred_on=fields.get("date", ""),
        amount=amount,
        confidence=Confidence(score=confidence_score, tier=_tier(confidence_score), method=parser),
        reason=f"extracted by deterministic {parser} parser",
        ingestion_batch_id=context.ingestion_batch_id,
        extraction_version_id=context.extraction_version_id,
        source_page=source_location.page_number,
        source_row=source_location.row_number,
        source_sheet=source_location.sheet_name,
    )
    return ExtractedRecord(
        organization_id=context.organization_id,
        source_document_id=context.source_document_id,
        ingestion_batch_id=context.ingestion_batch_id,
        extraction_version_id=context.extraction_version_id,
        source_record_id=source_record_id,
        record_type=record_type,
        fields=fields,
        field_confidences=field_confidences,
        source_location=source_location,
        evidence=evidence,
        confidence=confidence_score,
        warnings=tuple(sorted(set(warnings))),
        missing_fields=missing_fields,
        quality_flags=quality_flags,
    )


def _finalize_result(
    context: ExtractionContext,
    parser: str,
    records: list[ExtractedRecord],
    warnings: list[str] | None = None,
) -> ExtractionResult:
    warnings = list(warnings or [])
    seen: dict[tuple[str, str, str, str, str], int] = {}
    finalized: list[ExtractedRecord] = []
    for record in records:
        identity = (
            record.record_type,
            record.fields.get("date", ""),
            record.fields.get("reference", ""),
            record.fields.get("amount", ""),
            record.fields.get("currency", ""),
        )
        if identity in seen and any(identity[1:]):
            duplicate_confidence = min(record.confidence, 0.68)
            duplicate_flags = set(record.quality_flags)
            duplicate_flags.discard("complete")
            duplicate_flags.update({"duplicate-risk", "needs-review"})
            record = replace(
                record,
                confidence=duplicate_confidence,
                evidence=replace(
                    record.evidence,
                    confidence=Confidence(
                        score=duplicate_confidence,
                        tier=_tier(duplicate_confidence),
                        method=record.evidence.confidence.method,
                    ),
                ),
                warnings=tuple(sorted(set(record.warnings + ("duplicate-record",)))),
                quality_flags=tuple(sorted(duplicate_flags)),
            )
        seen[identity] = seen.get(identity, 0) + 1
        finalized.append(record)

    all_flags = {flag for record in finalized for flag in record.quality_flags}
    if not finalized:
        all_flags.update({"incomplete", "needs-review"})
        warnings.append("no-records-extracted")
    if all_flags == {"complete"}:
        quality_flags: tuple[str, ...] = ("complete",)
    else:
        all_flags.discard("complete")
        quality_flags = tuple(sorted(all_flags))
    result = ExtractionResult(
        context=context,
        parser=parser,
        schema_version=SCHEMA_VERSION,
        records=tuple(finalized),
        warnings=tuple(sorted(set(warnings))),
        quality_flags=quality_flags,
    )
    result.validate()
    return result


def _empty_result(context: ExtractionContext, parser: str, warning: str) -> ExtractionResult:
    result = ExtractionResult(
        context=context,
        parser=parser,
        schema_version=SCHEMA_VERSION,
        warnings=(warning,),
        quality_flags=("incomplete", "needs-review"),
    )
    result.validate()
    return result


def _quality_flags(confidence: float, missing_fields: tuple[str, ...], warnings: list[str]) -> tuple[str, ...]:
    flags: set[str] = set()
    if missing_fields:
        flags.add("incomplete")
    if confidence < 0.70:
        flags.add("low-confidence")
    if warnings or missing_fields or confidence < 0.70:
        flags.add("needs-review")
    if not flags:
        flags.add("complete")
    return tuple(sorted(flags))


def _infer_record_type(fields: dict[str, str], file_name: str) -> str:
    explicit = _slug(fields.get("type", ""))
    singular = explicit[:-1] if explicit.endswith("s") else explicit
    if singular in _RECORD_TYPES:
        return singular
    haystack = f"{file_name} {fields.get('description', '')}".lower()
    for candidate in ("receipt", "invoice", "bill", "contract", "premium", "claim", "payment", "transaction"):
        if candidate in haystack:
            return candidate
    return "transaction"


def _normalize_header(value: str) -> str:
    return _slug(value.strip().lower())


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def _normalize_date(value: str) -> tuple[str, str]:
    normalized = value.strip()
    try:
        return datetime.fromisoformat(normalized.replace("Z", "+00:00")).date().isoformat(), ""
    except ValueError:
        pass
    for date_format in (
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%B %d, %Y",
        "%b %d, %Y",
    ):
        try:
            parsed = datetime.strptime(normalized, date_format).date().isoformat()
            ambiguous = bool(re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", normalized))
            first, second, _ = normalized.split("/") if ambiguous else ("99", "99", "")
            warning = "ambiguous-date" if ambiguous and int(first) <= 12 and int(second) <= 12 else ""
            return parsed, warning
        except ValueError:
            continue
    return normalized, "invalid-date"


def _normalize_amount(value: str) -> tuple[str, str]:
    normalized = value.strip()
    negative = normalized.startswith("(") and normalized.endswith(")")
    normalized = re.sub(r"[^0-9,.-]", "", normalized)
    if normalized.count(",") > 0 and normalized.count(".") == 0:
        parts = normalized.split(",")
        if len(parts[-1]) in (1, 2) and len(parts) == 2:
            normalized = ".".join(parts)
        else:
            normalized = "".join(parts)
    else:
        normalized = normalized.replace(",", "")
    try:
        amount = Decimal(normalized)
    except InvalidOperation:
        return value.strip(), "invalid-amount"
    if negative:
        amount = -amount
    return format(amount, "f"), ""


def _money_from_fields(fields: dict[str, str]) -> Money | None:
    amount_value = fields.get("amount")
    if not amount_value:
        return None
    try:
        amount = Decimal(amount_value)
    except InvalidOperation:
        return None
    currency = fields.get("currency", "").upper()
    if not currency:
        return None
    precision = 0 if currency in {"RWF", "UGX", "JPY"} else 2
    minor_units = int(amount * (10**precision))
    return Money(currency=currency, minor_units=minor_units, precision=precision)


def _tier(score: float) -> str:
    if score >= 0.95:
        return "auto"
    if score >= 0.70:
        return "suggested"
    return "review"

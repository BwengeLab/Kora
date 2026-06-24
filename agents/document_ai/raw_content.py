from __future__ import annotations

import csv
import io
import json
import os
import subprocess
import tempfile
from collections import defaultdict
from dataclasses import replace
from pathlib import Path
from typing import Any

from agents.document_ai.enterprise_schemas import (
    BoundingBox,
    ContentCell,
    ContentLine,
    ContentPage,
    ContentTable,
    ContentToken,
    ExtractionQuality,
    RawDocumentContent,
)


def extract_raw_content(
    content: bytes,
    file_name: str,
    detected_type: str,
    language: str = "eng",
) -> RawDocumentContent:
    if detected_type == "csv":
        return _csv_content(content)
    if detected_type == "xlsx":
        return _excel_content(content)
    if detected_type == "pdf":
        return _pdf_content(content, language)
    if detected_type in {"png", "jpeg", "webp", "tiff"}:
        return _image_content(content, language)
    raise ValueError(f"raw extraction does not support {detected_type or Path(file_name).suffix}")


def _csv_content(content: bytes) -> RawDocumentContent:
    text = content.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect=dialect))
    lines = tuple(
        ContentLine(
            text="\t".join(row),
            reading_order=index,
            tokens=tuple(ContentToken(cell, 1.0) for cell in row if cell),
        )
        for index, row in enumerate(rows)
        if any(cell for cell in row)
    )
    column_count = max((len(row) for row in rows), default=0)
    cells = tuple(
        ContentCell(cell, row_index, column_index)
        for row_index, row in enumerate(rows)
        for column_index, cell in enumerate(row)
    )
    table = (
        ContentTable(1, 0, len(rows), column_count, cells, 1.0)
        if rows and column_count
        else None
    )
    return _document(
        (ContentPage(1, column_count, len(rows), lines, (table,) if table else ()),),
        "csv-structural",
    )


def _excel_content(content: bytes) -> RawDocumentContent:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    pages: list[ContentPage] = []
    for page_number, sheet in enumerate(workbook.worksheets, start=1):
        lines: list[ContentLine] = []
        for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = ["" if value is None else str(value) for value in values]
            if not any(cells):
                continue
            lines.append(
                ContentLine(
                    text="\t".join(cells),
                    reading_order=row_number - 1,
                    tokens=tuple(ContentToken(cell, 1.0) for cell in cells if cell),
                )
            )
        table_cells = tuple(
            ContentCell(
                "" if cell.value is None else str(cell.value),
                cell.row - 1,
                cell.column - 1,
            )
            for row in sheet.iter_rows()
            for cell in row
        )
        tables = (
            ContentTable(
                page_number,
                0,
                sheet.max_row,
                sheet.max_column,
                table_cells,
                1.0,
            ),
        ) if sheet.max_row and sheet.max_column else ()
        pages.append(
            ContentPage(page_number, sheet.max_column, sheet.max_row, tuple(lines), tables)
        )
    return _document(tuple(pages), "excel-structural")


def _pdf_content(content: bytes, language: str) -> RawDocumentContent:
    try:
        import pymupdf  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("PyMuPDF is required for layout-preserving PDF extraction") from exc
    document = pymupdf.open(stream=content, filetype="pdf")
    pages: list[ContentPage] = []
    methods: set[str] = set()
    try:
        for page_index in range(document.page_count):
            page: Any = document.load_page(page_index)
            page_number = page_index + 1
            width, height = int(page.rect.width), int(page.rect.height)
            words = page.get_text("words", sort=True)
            if words:
                pages.append(_pymupdf_page(words, page_number, width, height))
                methods.add("pdf-text-layout")
                continue
            pixmap = page.get_pixmap(matrix=pymupdf.Matrix(2, 2), alpha=False)
            image_bytes = pixmap.tobytes("png")
            image_page, method = _ocr_image_page(image_bytes, language, page_number)
            pages.append(image_page)
            methods.add(f"pdf-{method}")
    finally:
        document.close()
    return _document(
        tuple(pages), "+".join(sorted(methods)) or "pdf-empty", language
    )


def _pymupdf_page(
    words: list[tuple[Any, ...]], page_number: int, width: int, height: int
) -> ContentPage:
    grouped: dict[tuple[int, int], list[tuple[Any, ...]]] = defaultdict(list)
    for word in words:
        grouped[(int(word[5]), int(word[6]))].append(word)
    lines: list[ContentLine] = []
    for reading_order, key in enumerate(sorted(grouped)):
        row = sorted(grouped[key], key=lambda word: int(word[7]))
        tokens = tuple(
            ContentToken(
                text=str(word[4]),
                confidence=1.0,
                bounding_box=_box(page_number, word[0], word[1], word[2], word[3], width, height),
            )
            for word in row
        )
        lines.append(_line_from_tokens(tokens, reading_order))
    page = ContentPage(page_number, width, height, tuple(lines))
    return replace(page, tables=_detect_page_tables(page))


def _image_content(content: bytes, language: str) -> RawDocumentContent:
    page, method = _ocr_image_page(content, language, 1)
    return _document((page,), method, language)


def _ocr_image_page(
    content: bytes, language: str, page_number: int
) -> tuple[ContentPage, str]:
    from PIL import Image

    with Image.open(io.BytesIO(content)) as source:
        image = source.convert("RGB")
    if os.getenv("KORA_OCR_PREPROCESS", "").lower() == "true":
        image = _prepare_ocr_image(image)
    try:
        return _tesseract_page(image, language, page_number), "tesseract-layout"
    except Exception:
        if os.name != "nt":
            raise
        return _windows_page(image, page_number), "windows-ocr-layout-unscored"


def _prepare_ocr_image(image: Any) -> Any:
    from PIL import Image, ImageFilter

    longest_edge = max(image.size)
    scale = min(2.0, 3200 / longest_edge) if longest_edge else 1.0
    if scale > 1.05:
        image = image.resize(
            (round(image.width * scale), round(image.height * scale)),
            Image.Resampling.LANCZOS,
        )
    return image.filter(ImageFilter.SHARPEN)


def _tesseract_page(image: Any, language: str, page_number: int) -> ContentPage:
    import pytesseract  # type: ignore[import-untyped]

    data = pytesseract.image_to_data(
        image, lang=language, output_type=pytesseract.Output.DICT, config="--psm 3"
    )
    grouped: dict[tuple[int, int, int], list[ContentToken]] = defaultdict(list)
    width, height = image.size
    for index, raw_text in enumerate(data["text"]):
        text = str(raw_text).strip()
        confidence = float(data["conf"][index])
        if not text or confidence < 0:
            continue
        key = (
            int(data["block_num"][index]),
            int(data["par_num"][index]),
            int(data["line_num"][index]),
        )
        x, y = int(data["left"][index]), int(data["top"][index])
        w, h = int(data["width"][index]), int(data["height"][index])
        grouped[key].append(
            ContentToken(
                text,
                max(0.0, min(confidence / 100, 1.0)),
                _box(page_number, x, y, x + w, y + h, width, height),
            )
        )
    lines = tuple(
        _line_from_tokens(tuple(grouped[key]), order)
        for order, key in enumerate(sorted(grouped))
    )
    page = ContentPage(page_number, width, height, lines)
    return replace(page, tables=_detect_page_tables(page))


def _windows_page(image: Any, page_number: int) -> ContentPage:
    payload = _run_windows_ocr(image)
    width, height = image.size
    page = ContentPage(
        page_number,
        width,
        height,
        _group_positioned_lines(payload, page_number, width, height),
    )
    return replace(page, tables=_detect_page_tables(page))


def _group_positioned_lines(
    raw_lines: list[dict[str, Any]], page_number: int, width: int, height: int
) -> tuple[ContentLine, ...]:
    ordered = sorted(raw_lines, key=lambda line: (float(line["y"]), float(line["x"])))
    bands: list[list[dict[str, Any]]] = []
    for raw_line in ordered:
        words = raw_line.get("words", [])
        if not words:
            continue
        top = min(float(word["y"]) for word in words)
        bottom = max(float(word["y"]) + float(word["height"]) for word in words)
        center = (top + bottom) / 2
        if not bands:
            bands.append([raw_line])
            continue
        previous_words = [word for line in bands[-1] for word in line["words"]]
        previous_top = min(float(word["y"]) for word in previous_words)
        previous_bottom = max(
            float(word["y"]) + float(word["height"]) for word in previous_words
        )
        previous_center = (previous_top + previous_bottom) / 2
        typical_height = max(
            1.0,
            sum(float(word["height"]) for word in words + previous_words)
            / len(words + previous_words),
        )
        overlaps = min(bottom, previous_bottom) - max(top, previous_top) > 0
        if overlaps or abs(center - previous_center) <= typical_height * 0.55:
            bands[-1].append(raw_line)
        else:
            bands.append([raw_line])

    output: list[ContentLine] = []
    for reading_order, band in enumerate(bands):
        sorted_lines = sorted(band, key=lambda line: float(line["x"]))
        tokens = tuple(
            ContentToken(
                str(word["text"]),
                0.0,
                _box(
                    page_number,
                    word["x"],
                    word["y"],
                    float(word["x"]) + float(word["width"]),
                    float(word["y"]) + float(word["height"]),
                    width,
                    height,
                ),
            )
            for line in sorted_lines
            for word in line["words"]
        )
        content_line = _line_from_tokens(tokens, reading_order)
        output.append(
            ContentLine(
                text="\t".join(_positioned_line_text(line) for line in sorted_lines),
                reading_order=reading_order,
                tokens=content_line.tokens,
                bounding_box=content_line.bounding_box,
            )
        )
    return tuple(output)


def _positioned_line_text(raw_line: dict[str, Any]) -> str:
    words = sorted(raw_line.get("words", []), key=lambda word: float(word["x"]))
    if not words:
        return str(raw_line.get("text", ""))
    character_widths = [
        float(word["width"]) / max(len(str(word["text"])), 1) for word in words
    ]
    typical_character_width = sum(character_widths) / len(character_widths)
    parts = [str(words[0]["text"])]
    previous_right = float(words[0]["x"]) + float(words[0]["width"])
    for word in words[1:]:
        gap = float(word["x"]) - previous_right
        separator = "\t" if gap > max(8.0, typical_character_width * 1.5) else " "
        parts.extend((separator, str(word["text"])))
        previous_right = float(word["x"]) + float(word["width"])
    return "".join(parts)


def _run_windows_ocr(image: Any) -> list[dict[str, Any]]:
    script = r"""
$ErrorActionPreference='Stop'; Add-Type -AssemblyName System.Runtime.WindowsRuntime
$g=([System.WindowsRuntimeSystemExtensions].GetMethods()|Where-Object{$_.Name -eq 'AsTask' -and $_.IsGenericMethod -and $_.GetParameters().Count -eq 1})[0]
function A($o,[Type]$t){$x=$g.MakeGenericMethod($t).Invoke($null,@($o));$x.Wait()|Out-Null;$x.Result}
[Windows.Storage.StorageFile,Windows.Storage,ContentType=WindowsRuntime]|Out-Null
[Windows.Storage.Streams.IRandomAccessStream,Windows.Storage.Streams,ContentType=WindowsRuntime]|Out-Null
[Windows.Graphics.Imaging.BitmapDecoder,Windows.Graphics.Imaging,ContentType=WindowsRuntime]|Out-Null
[Windows.Graphics.Imaging.SoftwareBitmap,Windows.Graphics.Imaging,ContentType=WindowsRuntime]|Out-Null
[Windows.Media.Ocr.OcrEngine,Windows.Foundation,ContentType=WindowsRuntime]|Out-Null
$f=A ([Windows.Storage.StorageFile]::GetFileFromPathAsync($env:KORA_OCR_FILE)) ([Windows.Storage.StorageFile]);$s=A ($f.OpenAsync([Windows.Storage.FileAccessMode]::Read)) ([Windows.Storage.Streams.IRandomAccessStream]);$d=A ([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($s)) ([Windows.Graphics.Imaging.BitmapDecoder]);$b=A ($d.GetSoftwareBitmapAsync()) ([Windows.Graphics.Imaging.SoftwareBitmap]);$e=[Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages();$r=A ($e.RecognizeAsync($b)) ([Windows.Media.Ocr.OcrResult])
$r.Lines|ForEach-Object{$w=@($_.Words);$first=$w[0];[pscustomobject]@{text=$_.Text;x=$first.BoundingRect.X;y=$first.BoundingRect.Y;words=@($w|ForEach-Object{[pscustomobject]@{text=$_.Text;x=$_.BoundingRect.X;y=$_.BoundingRect.Y;width=$_.BoundingRect.Width;height=$_.BoundingRect.Height}})}}|ConvertTo-Json -Depth 5 -Compress
"""
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temporary:
        path = temporary.name
    try:
        image.save(path, format="PNG")
        environment = dict(os.environ)
        environment["KORA_OCR_FILE"] = path
        completed = subprocess.run(
            ["powershell.exe", "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True,
            check=True,
            encoding="utf-8",
            errors="replace",
            env=environment,
            timeout=45,
        )
        payload = json.loads(completed.stdout)
        return [payload] if isinstance(payload, dict) else payload
    finally:
        try:
            os.remove(path)
        except FileNotFoundError:
            pass


def _line_from_tokens(tokens: tuple[ContentToken, ...], reading_order: int) -> ContentLine:
    boxes = [token.bounding_box for token in tokens if token.bounding_box]
    box = None
    if boxes:
        first = boxes[0]
        box = BoundingBox(
            first.page_number,
            min(item.x0 for item in boxes),
            min(item.y0 for item in boxes),
            max(item.x1 for item in boxes),
            max(item.y1 for item in boxes),
        )
    return ContentLine(" ".join(token.text for token in tokens), reading_order, tokens, box)


def _document(
    pages: tuple[ContentPage, ...],
    method: str,
    language: str = "",
) -> RawDocumentContent:
    full_text = "\n\f\n".join(
        "\n".join(line.text for line in page.lines) for page in pages
    )
    tokens = [token for page in pages for line in page.lines for token in line.tokens]
    positioned = sum(token.bounding_box is not None for token in tokens)
    unscored = "unscored" in method
    scores = [] if unscored else [token.confidence for token in tokens]
    low_confidence = sum(score < 0.5 for score in scores)
    printable = sum(character.isprintable() or character.isspace() for character in full_text)
    quality = ExtractionQuality(
        token_count=len(tokens),
        positioned_token_count=positioned,
        scored_token_count=len(scores),
        low_confidence_token_count=low_confidence,
        mean_confidence=sum(scores) / len(scores) if scores else None,
        coordinate_coverage=positioned / len(tokens) if tokens else 0.0,
        printable_character_ratio=printable / len(full_text) if full_text else 0.0,
    )
    warnings: list[str] = []
    if not tokens:
        warnings.append("no-text-tokens")
    if unscored:
        warnings.append("ocr-confidence-unavailable")
    if scores and low_confidence / len(scores) > 0.1:
        warnings.append("low-ocr-confidence")
    if tokens and quality.coordinate_coverage < 0.95 and method not in {
        "csv-structural",
        "excel-structural",
    }:
        warnings.append("incomplete-coordinate-coverage")
    if full_text and quality.printable_character_ratio < 0.99:
        warnings.append("non-printable-content")
    result = RawDocumentContent(full_text, pages, method, language, tuple(warnings), quality)
    result.validate()
    return result


def _detect_page_tables(page: ContentPage) -> tuple[ContentTable, ...]:
    candidates: list[tuple[int, tuple[ContentCell, ...]]] = []
    for line_index, line in enumerate(page.lines):
        groups = _split_line_into_cells(line.tokens)
        if len(groups) < 3:
            continue
        cells = tuple(
            _content_cell(group, 0, column_index)
            for column_index, group in enumerate(groups)
        )
        candidates.append((line_index, cells))

    runs: list[list[tuple[int, tuple[ContentCell, ...]]]] = []
    for candidate in candidates:
        if not runs or candidate[0] - runs[-1][-1][0] > 2:
            runs.append([candidate])
        else:
            runs[-1].append(candidate)

    tables: list[ContentTable] = []
    for run in runs:
        if len(run) < 3:
            continue
        row_cells: list[ContentCell] = []
        for row_index, (_, cells) in enumerate(run):
            row_cells.extend(replace(cell, row_index=row_index) for cell in cells)
        column_count = max(cell.column_index for cell in row_cells) + 1
        boxes = [cell.bounding_box for cell in row_cells if cell.bounding_box]
        table_box = _union_boxes(boxes) if boxes else None
        scores = [
            token.confidence for cell in row_cells for token in cell.tokens if token.confidence > 0
        ]
        tables.append(
            ContentTable(
                page_number=page.page_number,
                reading_order=run[0][0],
                row_count=len(run),
                column_count=column_count,
                cells=tuple(row_cells),
                confidence=sum(scores) / len(scores) if scores else 0.0,
                bounding_box=table_box,
            )
        )
    return tuple(tables)


def _split_line_into_cells(
    tokens: tuple[ContentToken, ...],
) -> tuple[tuple[ContentToken, ...], ...]:
    positioned = sorted(
        (token for token in tokens if token.bounding_box),
        key=lambda token: token.bounding_box.x0 if token.bounding_box else 0,
    )
    if len(positioned) < 3:
        return ()
    character_widths = [
        (token.bounding_box.x1 - token.bounding_box.x0) / max(len(token.text), 1)
        for token in positioned
        if token.bounding_box
    ]
    typical_character_width = sum(character_widths) / len(character_widths)
    groups: list[list[ContentToken]] = [[]]
    previous_right: float | None = None
    for token in positioned:
        token_text = token.text.strip()
        if token_text in {"|", "¦"}:
            if groups[-1]:
                groups.append([])
            previous_right = None
            continue
        has_leading_border = token_text[:1] in {"|", "¦"}
        has_trailing_border = token_text[-1:] in {"|", "¦"}
        if has_leading_border and groups[-1]:
            groups.append([])
        if has_leading_border or has_trailing_border:
            token_text = token_text.strip("|¦")
            if not token_text:
                previous_right = None
                continue
            token = replace(token, text=token_text)
        box = token.bounding_box
        assert box is not None
        gap = box.x0 - previous_right if previous_right is not None else 0.0
        if groups[-1] and gap > max(0.012, typical_character_width * 2.2):
            groups.append([])
        groups[-1].append(token)
        previous_right = box.x1
        if has_trailing_border:
            groups.append([])
            previous_right = None
    return tuple(tuple(group) for group in groups if group)


def _content_cell(
    tokens: tuple[ContentToken, ...], row_index: int, column_index: int
) -> ContentCell:
    boxes = [token.bounding_box for token in tokens if token.bounding_box]
    return ContentCell(
        text=" ".join(token.text for token in tokens),
        row_index=row_index,
        column_index=column_index,
        tokens=tokens,
        bounding_box=_union_boxes(boxes) if boxes else None,
    )


def _union_boxes(boxes: list[BoundingBox]) -> BoundingBox:
    first = boxes[0]
    return BoundingBox(
        first.page_number,
        min(box.x0 for box in boxes),
        min(box.y0 for box in boxes),
        max(box.x1 for box in boxes),
        max(box.y1 for box in boxes),
    )


def _box(
    page_number: int,
    x0: float,
    y0: float,
    x1: float,
    y1: float,
    width: int,
    height: int,
) -> BoundingBox:
    return BoundingBox(
        page_number,
        max(0.0, min(float(x0) / width, 1.0)),
        max(0.0, min(float(y0) / height, 1.0)),
        max(0.0, min(float(x1) / width, 1.0)),
        max(0.0, min(float(y1) / height, 1.0)),
    )
